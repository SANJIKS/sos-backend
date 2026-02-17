"""
Экспорт данных для WS Dashboard
"""
import csv
import logging
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import PermissionDenied

from ..models import Donation, DonationTransaction, DonationCampaign

logger = logging.getLogger(__name__)


def export_donations_csv(request):
    """
    Экспорт пожертвований в CSV
    
    Доступен только для superadmin пользователей.
    
    Параметры запроса:
    - start_date: Дата начала (YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS)
    - end_date: Дата окончания (YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS)
    - status: Фильтр по статусу (опционально)
    - is_recurring: Фильтр по рекуррентным платежам (true/false, опционально)
    - campaign_id: Фильтр по кампании (опционально)
    """
    # Проверяем права доступа
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied. Only superadmin users can export donations.'}, status=403)
    
    try:
        # Парсим параметры
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        status_filter = request.GET.get('status')
        is_recurring_filter = request.GET.get('is_recurring')
        campaign_id = request.GET.get('campaign_id')
        
        # Создаем queryset
        queryset = Donation.objects.all()
        
        # Фильтр по дате
        if start_date_str:
            try:
                start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__gte=start_date)
            except ValueError:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                    queryset = queryset.filter(created_at__gte=start_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid start_date format'}, status=400)
        
        if end_date_str:
            try:
                end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__lte=end_date)
            except ValueError:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                    end_date = end_date + timedelta(days=1)  # Включаем весь день
                    queryset = queryset.filter(created_at__lt=end_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid end_date format'}, status=400)
        
        # Фильтр по статусу
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Фильтр по рекуррентным платежам
        if is_recurring_filter is not None:
            is_recurring = is_recurring_filter.lower() == 'true'
            queryset = queryset.filter(is_recurring=is_recurring)
        
        # Фильтр по кампании
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
        
        # Создаем CSV ответ
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="donations_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Заголовки
        headers = [
            'Order ID',
            'Donation Code',
            'Donation UUID',
            'Donor Full Name',
            'Donor Email',
            'Donor Phone',
            'Amount',
            'Currency',
            'Donation Type',
            'Payment Method',
            'Status',
            'Is Recurring',
            'Parent Order ID',
            'Recurring Active',
            'Subscription Status',
            'Campaign Name',
            'Donor Source',
            'Donor Comment',
            'Created At',
            'Payment Completed At',
            'Salesforce ID',
            'Salesforce Synced',
        ]
        writer.writerow(headers)
        
        # Данные
        for donation in queryset.select_related('campaign').prefetch_related('transactions'):
            # Получаем Order ID (parent_order_id или первый transaction_id)
            order_id = donation.parent_order_id
            if not order_id and donation.transactions.exists():
                order_id = donation.transactions.first().transaction_id
            
            row = [
                order_id or '',
                donation.donation_code,
                str(donation.uuid),
                donation.donor_full_name,
                donation.donor_email,
                donation.donor_phone,
                str(donation.amount),
                donation.currency,
                donation.get_donation_type_display(),
                donation.get_payment_method_display(),
                donation.get_status_display(),
                'Yes' if donation.is_recurring else 'No',
                donation.parent_order_id or '',
                'Yes' if donation.recurring_active else 'No',
                donation.get_subscription_status_display() if donation.subscription_status else '',
                donation.campaign.name if donation.campaign else '',
                donation.get_donor_source_display(),
                donation.donor_comment,
                donation.created_at.isoformat() if donation.created_at else '',
                donation.payment_completed_at.isoformat() if donation.payment_completed_at else '',
                donation.salesforce_id or '',
                'Yes' if donation.salesforce_synced else 'No',
            ]
            writer.writerow(row)
        
        return response
        
    except Exception as e:
        logger.error(f"Export CSV error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def export_donations_excel(request):
    """
    Экспорт пожертвований в Excel
    
    Доступен только для superadmin пользователей.
    
    Параметры запроса:
    - start_date: Дата начала (YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS)
    - end_date: Дата окончания (YYYY-MM-DD или YYYY-MM-DDTHH:MM:SS)
    - status: Фильтр по статусу (опционально)
    - is_recurring: Фильтр по рекуррентным платежам (true/false, опционально)
    - campaign_id: Фильтр по кампании (опционально)
    """
    # Проверяем права доступа
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied. Only superadmin users can export donations.'}, status=403)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({
            'error': 'openpyxl library is required for Excel export. Install it with: pip install openpyxl'
        }, status=500)
    
    try:
        # Парсим параметры (такие же как для CSV)
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        status_filter = request.GET.get('status')
        is_recurring_filter = request.GET.get('is_recurring')
        campaign_id = request.GET.get('campaign_id')
        
        # Создаем queryset (такая же логика как для CSV)
        queryset = Donation.objects.all()
        
        # Фильтр по дате
        if start_date_str:
            try:
                start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__gte=start_date)
            except ValueError:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                    queryset = queryset.filter(created_at__gte=start_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid start_date format'}, status=400)
        
        if end_date_str:
            try:
                end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__lte=end_date)
            except ValueError:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                    end_date = end_date + timedelta(days=1)
                    queryset = queryset.filter(created_at__lt=end_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid end_date format'}, status=400)
        
        # Фильтр по статусу
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Фильтр по рекуррентным платежам
        if is_recurring_filter is not None:
            is_recurring = is_recurring_filter.lower() == 'true'
            queryset = queryset.filter(is_recurring=is_recurring)
        
        # Фильтр по кампании
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Donations"
        
        # Заголовки
        headers = [
            'Order ID',
            'Donation Code',
            'Donation UUID',
            'Donor Full Name',
            'Donor Email',
            'Donor Phone',
            'Amount',
            'Currency',
            'Donation Type',
            'Payment Method',
            'Status',
            'Is Recurring',
            'Parent Order ID',
            'Recurring Active',
            'Subscription Status',
            'Campaign Name',
            'Donor Source',
            'Donor Comment',
            'Created At',
            'Payment Completed At',
            'Salesforce ID',
            'Salesforce Synced',
        ]
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Данные
        row_num = 2
        for donation in queryset.select_related('campaign').prefetch_related('transactions'):
            # Получаем Order ID
            order_id = donation.parent_order_id
            if not order_id and donation.transactions.exists():
                order_id = donation.transactions.first().transaction_id
            
            row_data = [
                order_id or '',
                donation.donation_code,
                str(donation.uuid),
                donation.donor_full_name,
                donation.donor_email,
                donation.donor_phone,
                float(donation.amount),
                donation.currency,
                donation.get_donation_type_display(),
                donation.get_payment_method_display(),
                donation.get_status_display(),
                'Yes' if donation.is_recurring else 'No',
                donation.parent_order_id or '',
                'Yes' if donation.recurring_active else 'No',
                donation.get_subscription_status_display() if donation.subscription_status else '',
                donation.campaign.name if donation.campaign else '',
                donation.get_donor_source_display(),
                donation.donor_comment,
                donation.created_at.isoformat() if donation.created_at else '',
                donation.payment_completed_at.isoformat() if donation.payment_completed_at else '',
                donation.salesforce_id or '',
                'Yes' if donation.salesforce_synced else 'No',
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
            
            row_num += 1
        
        # Автоматическая ширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        # Создаем HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="donations_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Export Excel error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_donations(request):
    """
    Универсальный endpoint для экспорта пожертвований
    
    Доступен только для superadmin пользователей.
    
    Параметры:
    - format: csv или xlsx (по умолчанию csv)
    - start_date: Дата начала (YYYY-MM-DD или ISO format)
    - end_date: Дата окончания (YYYY-MM-DD или ISO format)
    - status: Фильтр по статусу
    - is_recurring: Фильтр по рекуррентным платежам (true/false)
    - campaign_id: Фильтр по кампании
    """
    # Проверяем, что пользователь является superuser
    if not request.user.is_authenticated or not request.user.is_superuser:
        raise PermissionDenied("Only superadmin users can export donations")
    
    export_format = request.GET.get('format', 'csv').lower()
    
    if export_format == 'xlsx' or export_format == 'excel':
        return export_donations_excel(request)
    else:
        return export_donations_csv(request)
